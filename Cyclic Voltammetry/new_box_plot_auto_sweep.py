import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.chart import ScatterChart, Reference, Series
import datetime
import os
import sys
import struct
import time
import csv
import math
from datetime import datetime
from PySide2.QtWidgets import QApplication
from SquidstatPyLibrary import AisDeviceTracker
from SquidstatPyLibrary import AisCompRange
from SquidstatPyLibrary import AisDCData
from SquidstatPyLibrary import AisACData
from SquidstatPyLibrary import AisExperimentNode
from SquidstatPyLibrary import AisErrorCode
from SquidstatPyLibrary import AisExperiment
from SquidstatPyLibrary import AisInstrumentHandler

from SquidstatPyLibrary import AisCyclicVoltammetryElement
from SquidstatPyLibrary import AisConstantPotElement
from SquidstatPyLibrary import AisEISPotentiostaticElement
from SquidstatPyLibrary import AisConstantCurrentElement

# How to use this code:
    # Edit and save this code. Specifically, edit the SWEEPS variable to contain all the sweep rates you did for your device, in mV/s
        # Also change the DEVICE_BATCH, DEVICE_NUMBER, and ACTIVE_AREA variables
    # Attach the device you would like to test to the Squidstat probes
    # In your terminal/command line, navigate (using the cd command) to the folder containing this python file
    # In the terminal/command line type: python box_plot_auto_sweep.py
        # Press enter to run the code
        # As the code runs, a lot of data may be pasted into the console. This is the CV data, which is being saved to csv files
    # The code will output a new folder called "CV Experiment [timestamp]", where [timestamp] is the time at which this code was run.
        # The folder contains an Excel workbook called "Box Plot [timestamp]", along with csv files that hold the raw data for the sweep rates you input
        # The "Box Plot [timestamp]" workbook contains a Summary sheet, which shows your CV curves and a linear fit for finding capacitance/area
            # The remaining sheets contain the content of the previously mentioned csv files

# Enter your sweep rates in mV/s, in descending order
SWEEPS = [1500, 1250, 1000, 750, 500]

# Enter device information
DEVICE_BATCH = "Ru Devices 9_1_23"
DEVICE_NUMBER = "Ru 7.5cm2 Device 1 350C 20Ar Ramp"
ACTIVE_AREA = 7.5 # cm2
MAX_VOLTAGE = 1 # V
MIN_VOLTAGE = -1 # V

def fit_and_center(sheet):
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Center align cells with data
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(horizontal='center', vertical='center')

def on_experiment_completed(channel):
    print("Experiment Completed: %d" % channel)
    app.quit()

def save_data_to_csv(channel, data):
    # Append the data to the list
    data_list.append([1, "Filler", "Cyclic Voltammetry", data.timestamp, data.workingElectrodeVoltage, data.workingElectrodeVoltage, data.current, 0, 0, data.counterElectrodeVoltage])
    
    # Check if it's the first 10 rows and remove rows with WorkingElectrodeVoltage < 0.95
    if len(data_list) <= 10:
        data_list[:] = [row for row in data_list if math.isclose(abs(row[4]), 0.95, rel_tol=0.0, abs_tol=0.05)]

def run_cyclic_voltammetry_experiment(dEdt):
    # Clear the data list for each run
    data_list.clear()

    # Append the cyclic voltammetry element to the experiment and set it to run 1 time
    experiment = AisExperiment()
    experiment.appendElement(AisConstantPotElement(-1, 0.25, 0.5), 1) # Add a half second of quiet time before each CV experiment to get rid of tail
    experiment.appendElement(cvElement, 1)

    handler.uploadExperimentToChannel(0, experiment)
    handler.startUploadedExperiment(0)

    app.exec_()

    # Write data to a CSV file for this run within the created folder
    csv_filename = f'{folder_name}/{int(dEdt*1000)} mV_s.csv'
    with open(csv_filename, mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(['Repeats', 'Filler', 'Step name', 'Elapsed Time (s)', 'Working Electrode (V)', 'Working Electrode vs. NHE (V)', 'Current (A)', 'Current Density (A/m^2)', 'Cumulative Charge (mAh)', 'Counter Electrode (V)'])  # Write header
        writer.writerows(data_list)  # Write the data

##################################################################################################################################################################################################################

app = QApplication()
data_list = []

tracker = AisDeviceTracker.Instance()

tracker.newDeviceConnected.connect(lambda deviceName: print("Device is Connected: %s" % deviceName))
tracker.connectToDeviceOnComPort("COM17")

handler = tracker.getInstrumentHandler("Plus1931")

print(tracker.getConnectedDevices())

# handler.activeDCDataReady.connect(lambda channel, data: print("timestamp:", "{:.9f}".format(data.timestamp), "workingElectrodeVoltage: ", "{:.9f}".format(data.workingElectrodeVoltage), "current: ", "{:.9f}".format(data.current)))
handler.activeDCDataReady.connect(save_data_to_csv)
handler.experimentNewElementStarting.connect(lambda channel, data: print("New Node beginning:", data.stepName, "step number: ", data.stepNumber, " step sub : ", data.substepNumber))
handler.experimentStopped.connect(on_experiment_completed)

# # Initialize the CyclicVoltammetryElement with initial dEdt value
# cvElement = AisCyclicVoltammetryElement(-1, 1, -1, -1, 0.5, 0.0024)

# List of dEdt values to run experiments with
dEdt_values = [sweep/1000 for sweep in SWEEPS]

# Get the current timestamp to create a folder name
timestamp_str = datetime.now().strftime("%Y%m%d %H_%M_%S")

# Create a folder with the timestamp and "CV Experiment" in the name
folder_name = f"CV Experiment {timestamp_str}"
os.makedirs(folder_name, exist_ok=True)

# Run 5 cyclic voltammetry experiments with different dEdt values
for dEdt in dEdt_values:

    if dEdt == dEdt_values[0]:
        cvElement = AisCyclicVoltammetryElement(-1, 1, -1, -1, dEdt, 0.0024)
        run_cyclic_voltammetry_experiment(dEdt)

        cvElement = AisCyclicVoltammetryElement(-1, 1, -1, -1, dEdt, 0.0024)
    else:
        cvElement = AisCyclicVoltammetryElement(-1, 1, -1, -1, dEdt, 0.0024)

    run_cyclic_voltammetry_experiment(dEdt)

app.quit()

##################################################################################################################################################################################################################

# Define the file paths
xlsx_file = "New Box Plot Sheet Template.xlsx"
csv_files = [f"{sweep} mV_s.csv" for sweep in SWEEPS]

# Create a new workbook
workbook = Workbook()

# Read the XLSX file and create a new sheet in the workbook
xlsx_workbook = load_workbook(xlsx_file)
xlsx_data = pd.DataFrame(xlsx_workbook[xlsx_workbook.sheetnames[0]].values)
xlsx_sheet = workbook.active
xlsx_sheet.title = "Summary"
for row in dataframe_to_rows(xlsx_data, index=False, header=False):
    for cell in row:
        if cell is str and cell.startswith("="):
            cell = "'" + cell
    xlsx_sheet.append(row)

# Put user entered sweep rates (from SWEEPS list) into Summary sheet
for i in range(len(SWEEPS)):
    xlsx_sheet.cell(row=(9+i), column=1).value = SWEEPS[i]

# Put user entered device information into the Summary sheet
xlsx_sheet.cell(row=1, column=2).value = DEVICE_BATCH
xlsx_sheet.cell(row=2, column=2).value = DEVICE_NUMBER
xlsx_sheet.cell(row=3, column=2).value = ACTIVE_AREA
xlsx_sheet.cell(row=4, column=2).value = MAX_VOLTAGE
xlsx_sheet.cell(row=5, column=2).value = MIN_VOLTAGE

box_plot = ScatterChart()
box_plot.height = 13.5
box_plot.width = 21.5

# Read the CSV files and create a new sheet in the workbook for each file
for i, csv_file in enumerate(csv_files, start=1):
    csv_data = pd.read_csv(folder_name + "/" + csv_file)
    csv_sheet = workbook.create_sheet(title=f"{csv_files[i-1].split('.')[0]}")
    for row in dataframe_to_rows(csv_data, index=False, header=True):
        csv_sheet.append(row)

    # Bold the header cells
    if len(csv_data.columns) > 0:
        for cell in csv_sheet[1]:
            cell.font = Font(bold=True)

    fit_and_center(csv_sheet)

    csv_sheet.cell(row=1, column=8).value = 'Current/Area (A/cm2)' # change header for this column (it defaults as A/m2 instead of A/cm2)

    # Modify values in column H to get correct current/area
    row_number = 2  # Start row number after header
    for row in csv_sheet.iter_rows(min_row=2, max_col=8, max_row=csv_sheet.max_row, values_only=True):
        csv_sheet.cell(row=row_number, column=8).value = f"=G{row_number}/Summary!$B$3"
        row_number += 1

    # Get the row number for zero voltage current and put the zero voltage current densities for each sweep rates into an Excel table
    filtered_csv_data = csv_data[csv_data['Current (A)'] > 0]
    filtered_csv_data = filtered_csv_data['Working Electrode (V)']
    min_abs_value = filtered_csv_data.abs().idxmin()
    # xlsx_sheet.cell(row=(9+i-1), column=2).value = f"=\'{SWEEPS[i-1]} mV_s\'!" + csv_sheet.cell(row=(min_abs_value+2), column=8).value[1:]

    # Add series for each csv sheet to the scatter plot
    x_values = Reference(csv_sheet, min_col=5, min_row=3, max_row=csv_sheet.max_row)
    y_values = Reference(csv_sheet, min_col=7, min_row=3, max_row=csv_sheet.max_row)
    series = Series(y_values, x_values, title=f'{SWEEPS[i-1]} mV/s')
    box_plot.series.append(series)

xlsx_sheet.cell(row=9, column=2).value = "=XLOOKUP(0,ABS('1500 mV_s'!E3:'1500 mV_s'!E700),'1500 mV_s'!G3:'1500 mV_s'!G700,,1)"
xlsx_sheet.cell(row=10, column=2).value = "=XLOOKUP(0,ABS('1250 mV_s'!E3:'1250 mV_s'!E700),'1250 mV_s'!G3:'1250 mV_s'!G700,,1)"
xlsx_sheet.cell(row=11, column=2).value = "=XLOOKUP(0,ABS('1000 mV_s'!E3:'1000 mV_s'!E700),'1000 mV_s'!G3:'1000 mV_s'!G700,,1)"
xlsx_sheet.cell(row=12, column=2).value = "=XLOOKUP(0,ABS('750 mV_s'!E3:'750 mV_s'!E700),'750 mV_s'!G3:'750 mV_s'!G700,,1)"
xlsx_sheet.cell(row=13, column=2).value = "=XLOOKUP(0,ABS('500 mV_s'!E3:'500 mV_s'!E1200),'500 mV_s'!G3:'500 mV_s'!G1200,,1)"

xlsx_sheet.cell(row=9, column=3).value = "=SUM(U4:U3500)/($B$4-$B$5)/2"
xlsx_sheet.cell(row=10, column=3).value = "=SUM(V4:V3500)/($B$4-$B$5)/2"
xlsx_sheet.cell(row=11, column=3).value = "=SUM(W4:W3500)/($B$4-$B$5)/2"
xlsx_sheet.cell(row=12, column=3).value = "=SUM(X4:X3500)/($B$4-$B$5)/2"
xlsx_sheet.cell(row=13, column=3).value = "=SUM(Y4:Y3500)/($B$4-$B$5)/2"

xlsx_sheet.cell(row=15, column=2).value = "=INDEX(LINEST(B9:B13,A9:A13/1000,TRUE,TRUE),1,1)*10^9"
xlsx_sheet.cell(row=16, column=2).value = "=INDEX(LINEST(B9:B13,A9:A13/1000,TRUE,TRUE),2,1)*10^9"
xlsx_sheet.cell(row=17, column=2).value = "=B15/$B$3"
xlsx_sheet.cell(row=18, column=2).value = "=INDEX(LINEST(B9:B13,A9:A13/1000,TRUE,TRUE),1,2)*10^9"
xlsx_sheet.cell(row=19, column=2).value = "=INDEX(LINEST(B9:B13,A9:A13/1000,TRUE,TRUE),2,2)*10^9"

xlsx_sheet.cell(row=15, column=3).value = "=INDEX(LINEST(C9:C13,A9:A13/1000,TRUE,TRUE),1,1)*10^9"
xlsx_sheet.cell(row=16, column=3).value = "=INDEX(LINEST(C9:C13,A9:A13/1000,TRUE,TRUE),2,1)*10^9"
xlsx_sheet.cell(row=17, column=3).value = "=C15/$B$3"
xlsx_sheet.cell(row=18, column=3).value = "=INDEX(LINEST(C9:C13,A9:A13/1000,TRUE,TRUE),1,2)*10^9"
xlsx_sheet.cell(row=19, column=3).value = "=INDEX(LINEST(C9:C13,A9:A13/1000,TRUE,TRUE),2,2)*10^9"

xlsx_sheet.cell(row=37, column=2).value = "=(0.5--0.5)/(XLOOKUP(0.5,'1500 mV_s'!E2:E700,'1500 mV_s'!G2:G700,,-1)-XLOOKUP(-0.5,'1500 mV_s'!E2:E700,'1500 mV_s'!G2:G700,,-1))/10^6"
xlsx_sheet.cell(row=38, column=2).value = "=(0.5--0.5)/(XLOOKUP(0.5,'1250 mV_s'!E2:E700,'1250 mV_s'!G2:G700,,-1)-XLOOKUP(-0.5,'1250 mV_s'!E2:E700,'1250 mV_s'!G2:G700,,-1))/10^6"
xlsx_sheet.cell(row=39, column=2).value = "=(0.5--0.5)/(XLOOKUP(0.5,'1000 mV_s'!E2:E700,'1000 mV_s'!G2:G700,,-1)-XLOOKUP(-0.5,'1000 mV_s'!E2:E700,'1000 mV_s'!G2:G700,,-1))/10^6"
xlsx_sheet.cell(row=40, column=2).value = "=(0.5--0.5)/(XLOOKUP(0.5,'750 mV_s'!E2:E900,'750 mV_s'!G2:G900,,-1)-XLOOKUP(-0.5,'750 mV_s'!E2:E900,'750 mV_s'!G2:G900,,-1))/10^6"
xlsx_sheet.cell(row=41, column=2).value = "=(0.5--0.5)/(XLOOKUP(0.5,'500 mV_s'!E2:E1400,'500 mV_s'!G2:G1400,,-1)-XLOOKUP(-0.5,'500 mV_s'!E2:E1400,'500 mV_s'!G2:G1400,,-1))/10^6"

xlsx_sheet.cell(row=43, column=2).value = "=AVERAGE(B37:B41)"
xlsx_sheet.cell(row=44, column=2).value = "=STDEV.S(B37:B41)/SQRT(COUNT(A37:A41))"

# Add the chart to the worksheet
# box_plot.series = box_plot.series[::-1] # reverse the order of the series so that the plot legend is in descending order of sweep rates
xlsx_sheet.add_chart(box_plot, "E1")

# Adjust column widths and center alignment in Summary sheet
fit_and_center(xlsx_sheet)

# Add a scatter plot for zero voltage current density vs sweep rate
linear_plot = ScatterChart()
x_values = Reference(xlsx_sheet, min_col=1, min_row=37, max_row=(37+len(SWEEPS)-1))
y_values = Reference(xlsx_sheet, min_col=2, min_row=9, max_row=(9+len(SWEEPS)-1))
series = Series(y_values, x_values, title='Zero Voltage Current vs Sweep Rate')
linear_plot.series.append(series)
linear_plot.height = 13.5
linear_plot.width = 21.5
xlsx_sheet.add_chart(linear_plot, "E27")

# Save the combined data as a new XLSX file
combined_file = "Box Plot " + timestamp_str + ".xlsx"
workbook.save(folder_name + "/" + combined_file)

print(f"Data combined and saved successfully at {folder_name + '/' + combined_file}.")