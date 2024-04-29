import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.chart import ScatterChart, Reference, Series
import os
from datetime import datetime
import glob

# How to use this code:
    # Edit and save this code. Specifically, edit the SWEEPS variable to contain all the sweep rates you did for your device, in mV/s
        # Also change the DEVICE_BATCH, DEVICE_NUMBER, and ACTIVE_AREA variables
    # In your file browser, put this file in the same folder as your cyclic voltammetry folders recorded from your Squidstat
    # Make sure each folder and the csv data file it contains start with the sweep rate for that CV, in mV/s
        # For instance, if you ran a CV at 1V/s, make sure the csv file sits at something like: 1000 mV_s Cyclic Voltammetry Plus1930 ch1 (2024-04-29 09_52_08)/2_1000 mV_s Cyclic Voltammetry Plus1930 ch1 (2024-04-29 09_52_08) 20240429 095213
            # The 2_ prefix on the csv file is because there was a quiet time during the start of the CV experiment. CTRL + f this document for "prefixes =" to find the line to change if you didn't have a quiet time/there is no 2_ prefix.
    # In your terminal/command line, navigate (using the cd command) to the folder containing this python file
    # In the terminal/command line type: python analyze_box_plots.py
        # Press enter to run the code
        # The code will not run if you don't have python installed. You can install python from the Microsoft Store, python.org/downloads/, or anaconda.com/download
        # If you get an error that you don't have any of the imported packages (such as openpyxl), install them using: "pip install [package name]" in the command line
    # Once run, the code will output a new Excel workbook called "Box Plot [timestamp]"
        # The "Box Plot [timestamp]" workbook contains a Summary sheet, which shows your CV curves and a linear fit for finding capacitance/area
        # The remaining sheets contain the content of your csv files

################################################ Constants/User Action Required ################################################

# Enter your sweep rates in mV/s
SWEEPS = [1500, 1250, 1000, 750, 500]

# Enter device information
DEVICE_BATCH = "Amber Pt/ITO/SiO2"
DEVICE_NUMBER = "3"
ACTIVE_AREA = 1 # cm2
MAX_VOLTAGE = 1 # V
MIN_VOLTAGE = -1 # V

################################################ Functions ################################################

def find_files(prefixes):
    file_list = []
    for prefix in prefixes:
        files = glob.glob(f"**/{prefix}*.csv", recursive=True)
        if files:
            file_list.append(files[0])
    return file_list

def fit_and_center(sheet):
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Center align cells with data
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(horizontal='center', vertical='center')

def current_voltage_integral(workbook, worksheet_name, start_row, end_row):
    ws = workbook[worksheet_name]

    total_sum = 0

    for row in range(start_row, end_row + 1):
        cell_value = ws[f'E{row}'].value
        if cell_value is not None:
            if row > start_row:
                prev_cell_value = ws[f'E{row - 1}'].value
                total_sum += 0.5 * (cell_value - prev_cell_value) * (ws[f'G{row}'].value + ws[f'G{row - 1}'].value)

    return total_sum

################################################ Making and Populating the Excel Workbook ################################################

# Define the file paths
prefixes = ["2_" + str(sweep) for sweep in SWEEPS] # Assumes there is a 1_ prefixed file from a quiet time measurement. If this is not the case, change this line to: prefixes = [str(sweep) for sweep in SWEEPS]
csv_files = find_files(prefixes)

# Create a new workbook
workbook = Workbook()
xlsx_sheet = workbook.active
xlsx_sheet.title = "Summary"

# Put user entered sweep rates (from SWEEPS list) into Summary sheet, as well as column labels
for i in range(len(SWEEPS)):
    xlsx_sheet.cell(row=(9+i), column=1).value = SWEEPS[i]

xlsx_sheet.cell(row=8, column=1).value = "Sweep Rate (mV/s)"
xlsx_sheet.cell(row=8, column=2).value = "Zero Voltage Current (A)"
xlsx_sheet.cell(row=8, column=3).value = "Average Current (A)"

# Put user entered device information into the Summary sheet
xlsx_sheet.cell(row=1, column=1).value = "Device Batch"
xlsx_sheet.cell(row=1, column=2).value = DEVICE_BATCH

xlsx_sheet.cell(row=2, column=1).value = "Device Number"
xlsx_sheet.cell(row=2, column=2).value = DEVICE_NUMBER

xlsx_sheet.cell(row=3, column=1).value = "Active Area (cm2)"
xlsx_sheet.cell(row=3, column=2).value = ACTIVE_AREA

xlsx_sheet.cell(row=4, column=1).value = "Max Voltage (V)"
xlsx_sheet.cell(row=4, column=2).value = MAX_VOLTAGE

xlsx_sheet.cell(row=5, column=1).value = "Min Voltage (V)"
xlsx_sheet.cell(row=5, column=2).value = MIN_VOLTAGE

# Prepare Plot
box_plot = ScatterChart()
box_plot.height = 13.5
box_plot.width = 21.5

# Read the CSV files and create a new sheet in the workbook for each file
for i, csv_file in enumerate(csv_files, start=1):
    csv_data = pd.read_csv(csv_file)
    csv_sheet = workbook.create_sheet(title=f"{SWEEPS[i-1]} mV_s")
    for row in dataframe_to_rows(csv_data, index=False, header=True):
        csv_sheet.append(row)

    # Bold the header cells
    if len(csv_data.columns) > 0:
        for cell in csv_sheet[1]:
            cell.font = Font(bold=True)

    fit_and_center(csv_sheet)

    csv_sheet.cell(row=1, column=8).value = 'Current/Area (A/cm2)' # change header for this column (it defaults as A/m2 instead of A/cm2)

    # Modify values in column H to get correct current/area
    row_number = 2  # Start row number after header
    for row in csv_sheet.iter_rows(min_row=2, max_col=8, max_row=csv_sheet.max_row, values_only=True):
        csv_sheet.cell(row=row_number, column=8).value = f"=G{row_number}/Summary!$B$3"
        row_number += 1

    # Get the row number for zero voltage current and put the zero voltage current densities for each sweep rates into an Excel table
    filtered_csv_data = csv_data[csv_data['Current (A)'] > 0]
    filtered_csv_data = filtered_csv_data['Working Electrode (V)']
    min_abs_value = filtered_csv_data.abs().idxmin()

    # Add series for each csv sheet to the scatter plot
    x_values = Reference(csv_sheet, min_col=5, min_row=3, max_row=csv_sheet.max_row)
    y_values = Reference(csv_sheet, min_col=7, min_row=3, max_row=csv_sheet.max_row)
    series = Series(y_values, x_values, title=f'{SWEEPS[i-1]} mV/s')
    box_plot.series.append(series)

# Calculate the zero voltage current and average current
for row, sweep in enumerate(SWEEPS, 9):
    if sweep >= 750:
        xlsx_sheet.cell(row=row, column=2).value = f"=XLOOKUP(0,ABS('{sweep} mV_s'!E3:'{sweep} mV_s'!E700),'{sweep} mV_s'!G3:'{sweep} mV_s'!G700,,1)"
    else:
        xlsx_sheet.cell(row=row, column=2).value = f"=XLOOKUP(0,ABS('{sweep} mV_s'!E3:'{sweep} mV_s'!E1200),'{sweep} mV_s'!G3:'{sweep} mV_s'!G1200,,1)"

    integral = current_voltage_integral(workbook, f"{sweep} mV_s", 2, 3500)
    xlsx_sheet.cell(row=row, column=3).value = f"={integral}/($B$4-$B$5)/2"

# Calculate capacitance and stats based on zero voltage current
stats_stop_row = 9+len(SWEEPS)-1
xlsx_sheet.cell(row=20, column=2).value = f"=INDEX(LINEST(B9:B{stats_stop_row},A9:A{stats_stop_row}/1000,TRUE,TRUE),1,1)*10^9"
xlsx_sheet.cell(row=21, column=2).value = f"=INDEX(LINEST(B9:B{stats_stop_row},A9:A{stats_stop_row}/1000,TRUE,TRUE),2,1)*10^9"
xlsx_sheet.cell(row=22, column=2).value = "=B20/$B$3"
xlsx_sheet.cell(row=23, column=2).value = f"=INDEX(LINEST(B9:B{stats_stop_row},A9:A{stats_stop_row}/1000,TRUE,TRUE),1,2)*10^9"
xlsx_sheet.cell(row=24, column=2).value = f"=INDEX(LINEST(B9:B{stats_stop_row},A9:A{stats_stop_row}/1000,TRUE,TRUE),2,2)*10^9"

# Calculate capacitance and stats based on average current
xlsx_sheet.cell(row=20, column=3).value = f"=INDEX(LINEST(C9:C{stats_stop_row},A9:A{stats_stop_row}/1000,TRUE,TRUE),1,1)*10^9"
xlsx_sheet.cell(row=21, column=3).value = f"=INDEX(LINEST(C9:C{stats_stop_row},A9:A{stats_stop_row}/1000,TRUE,TRUE),2,1)*10^9"
xlsx_sheet.cell(row=22, column=3).value = f"=C20/$B$3"
xlsx_sheet.cell(row=23, column=3).value = f"=INDEX(LINEST(C9:C{stats_stop_row},A9:A{stats_stop_row}/1000,TRUE,TRUE),1,2)*10^9"
xlsx_sheet.cell(row=24, column=3).value = f"=INDEX(LINEST(C9:C{stats_stop_row},A9:A{stats_stop_row}/1000,TRUE,TRUE),2,2)*10^9"

# Row names for the above two blocks of code
xlsx_sheet.cell(row=20, column=1).value = "Capacitance (nF)"
xlsx_sheet.cell(row=21, column=1).value = "Std Error on Capacitance (nF)"
xlsx_sheet.cell(row=22, column=1).value = "Normalized Capacitance (nF/cm2)"
xlsx_sheet.cell(row=23, column=1).value = "Unaccounted for Current (nA)"
xlsx_sheet.cell(row=24, column=1).value = "Std Error on Current (nA)"

# Calculate average resistance in -0.5 V to 0.5 V range (Mohm). Also put in column headers
for row, sweep in enumerate(SWEEPS, 37):
    xlsx_sheet.cell(row=row, column=1).value = sweep/1000
    if sweep > 750:
        xlsx_sheet.cell(row=row, column=2).value = f"=(0.5--0.5)/(XLOOKUP(0.5,'{sweep} mV_s'!E2:E700,'{sweep} mV_s'!G2:G700,,-1)-XLOOKUP(-0.5,'{sweep} mV_s'!E2:E700,'{sweep} mV_s'!G2:G700,,-1))/10^6"
    elif sweep > 500:
        xlsx_sheet.cell(row=row, column=2).value = f"=(0.5--0.5)/(XLOOKUP(0.5,'{sweep} mV_s'!E2:E900,'{sweep} mV_s'!G2:G900,,-1)-XLOOKUP(-0.5,'{sweep} mV_s'!E2:E900,'{sweep} mV_s'!G2:G900,,-1))/10^6"
    else:
        xlsx_sheet.cell(row=row, column=2).value = f"=(0.5--0.5)/(XLOOKUP(0.5,'{sweep} mV_s'!E2:E1400,'{sweep} mV_s'!G2:G1400,,-1)-XLOOKUP(-0.5,'{sweep} mV_s'!E2:E1400,'{sweep} mV_s'!G2:G1400,,-1))/10^6"

xlsx_sheet.cell(row=36, column=1).value = "Sweep Rate (V/s)"
xlsx_sheet.cell(row=36, column=2).value = "Avg R in [-0.5,0.5]V (MΩ)"

# Get leakage resistance and std error for it
other_stats_stop_row = 37+len(SWEEPS)-1
xlsx_sheet.cell(row=other_stats_stop_row+2, column=1).value = "Leakage R (MΩ)"
xlsx_sheet.cell(row=other_stats_stop_row+2, column=2).value = f"=AVERAGE(B37:B{other_stats_stop_row})"
xlsx_sheet.cell(row=other_stats_stop_row+3, column=1).value = "Std Error on R (MΩ)"
xlsx_sheet.cell(row=other_stats_stop_row+3, column=2).value = f"=STDEV.S(B37:B{other_stats_stop_row})/SQRT(COUNT(A37:A{other_stats_stop_row}))"

################################################ Plots and Saving File ################################################

# Add the chart to the worksheet
xlsx_sheet.add_chart(box_plot, "E1")

# Adjust column widths and center alignment in Summary sheet
fit_and_center(xlsx_sheet)

# Add a scatter plot for zero voltage current density vs sweep rate
linear_plot = ScatterChart()
x_values = Reference(xlsx_sheet, min_col=1, min_row=37, max_row=(37+len(SWEEPS)-1))
y_values = Reference(xlsx_sheet, min_col=2, min_row=9, max_row=(9+len(SWEEPS)-1))
series = Series(y_values, x_values, title='Zero Voltage Current vs Sweep Rate')
linear_plot.series.append(series)
linear_plot.height = 13.5
linear_plot.width = 21.5
xlsx_sheet.add_chart(linear_plot, "E27")

# Save the combined data as a new XLSX file
timestamp_str = datetime.now().strftime("%Y%m%d %H_%M_%S")
combined_file = "Box Plot " + timestamp_str + ".xlsx"
workbook.save(combined_file)

print(f"Data combined and saved successfully at {combined_file}.")